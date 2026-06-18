import argparse
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from app.database import SessionLocal, init_db
from app.models import Constituent


DEFAULT_SUBGROUP = "Orange Active Voters"
LOCAL_WARD = "South"
STATUS_LABELS = {
    "A": "Active",
    "I": "Inactive",
}


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def title_clean(value) -> str:
    return clean(value).title()


def date_text(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return clean(value)


def full_name(row: dict) -> str:
    parts = [
        title_clean(row.get("First Name")),
        title_clean(row.get("Middle Name")),
        title_clean(row.get("Last Name")),
        title_clean(row.get("Suffix")),
    ]
    return " ".join(part for part in parts if part)


def status_label(value) -> str:
    raw = clean(value).upper()
    return STATUS_LABELS.get(raw, raw.title())


def ward_label(value) -> str:
    return title_clean(value) or "Unknown"


def note_for(row: dict, ward: str) -> str:
    local_note = "Local South Ward constituent." if ward.lower() == LOCAL_WARD.lower() else f"Citywide registered voter outside local South Ward; ward marker: {ward} Ward."
    details = [
        local_note,
        f"Party: {clean(row.get('Party')).upper() or 'Unknown'}",
        f"District: {clean(row.get('District')) or 'Unknown'}",
        f"DOB: {date_text(row.get('DOB')) or 'Unknown'}",
        f"Registered: {date_text(row.get('Reg Date')) or 'Unknown'}",
        f"Municipality: {clean(row.get('Municipality')) or 'City of Orange Township'}",
    ]
    return " | ".join(details)


def iter_voter_rows(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = list(next(rows))
    for values in rows:
        row = dict(zip(headers, values))
        voter_id = clean(row.get("ID"))
        if not voter_id or voter_id.startswith("Section -"):
            yield None
            continue
        yield row


def import_xlsx(path: Path, subgroup: str) -> dict[str, int | dict[str, int]]:
    init_db()
    db = SessionLocal()
    created = 0
    updated = 0
    skipped = 0
    by_ward: dict[str, int] = {}
    try:
        for row in iter_voter_rows(path):
            if row is None:
                skipped += 1
                continue

            voter_id = clean(row.get("ID"))
            ward = ward_label(row.get("Ward"))
            by_ward[ward] = by_ward.get(ward, 0) + 1
            local = ward.lower() == LOCAL_WARD.lower()
            street_no = clean(row.get("Street No."))
            zip_code = clean(row.get("Residence Zip"))

            payload = {
                "voter_id": voter_id,
                "first_name": title_clean(row.get("First Name")),
                "last_name": title_clean(row.get("Last Name")),
                "full_name": full_name(row),
                "street_no": street_no,
                "street": title_clean(row.get("Street Name")),
                "apt": clean(row.get("APT/UNIT")),
                "city": title_clean(row.get("Residence City")),
                "state": clean(row.get("Residence State")).upper(),
                "zip_code": zip_code.zfill(5) if zip_code.isdigit() else zip_code,
                "ward": ward,
                "subgroup": f"{subgroup} - South Ward" if local else subgroup,
                "voter_status": status_label(row.get("Status")),
                "mailin_request_date": None,
                "mailin_sent_date": None,
                "mailin_received_date": None,
                "days_to_return": None,
                "source_file": path.name,
                "notes": note_for(row, ward),
            }

            existing = db.query(Constituent).filter(Constituent.voter_id == voter_id).first()
            if existing:
                for key, value in payload.items():
                    setattr(existing, key, value)
                updated += 1
            else:
                db.add(Constituent(**payload))
                created += 1

        db.commit()
        return {"created": created, "updated": updated, "skipped": skipped, "by_ward": dict(sorted(by_ward.items()))}
    finally:
        db.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="Import Orange citywide active voter records into WardOS constituents.")
    parser.add_argument("xlsx_path", type=Path)
    parser.add_argument("--subgroup", default=DEFAULT_SUBGROUP)
    args = parser.parse_args()
    result = import_xlsx(args.xlsx_path, args.subgroup)
    print(result)


if __name__ == "__main__":
    main()
